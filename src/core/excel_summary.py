"""费用汇总 Excel 生成模块

从已处理的 PDF 发票中提取费用数据，按日期归集生成 Excel 汇总表。
"""
import os
import re
import logging
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── 费用类别关键字 ──────────────────────────────────────
_HOTEL_KEYWORDS = ['住宿', '酒店', '宾馆', '客房', '入住', '房费', '住宿费']
_TRANSPORT_KEYWORDS = [
    '高铁', '乘车', '滴滴', '通行费', '出行', '交通',
    '火车', '高铁票', '行程单',
]

# ── 日期提取模式 ────────────────────────────────────────
# 优先匹配带特定前缀的日期字段，fallback 到通用日期
_DATE_PATTERNS = [
    # 特定前缀日期（乘车/入住/开票/行程）
    (r'(?:乘车|行程|入住|开票|开具|消费|记账|发票)'
     r'日期\s*[:：]?\s*(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})[日]?'),
    # 通用日期（作为 fallback）
    (r'(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})[日]?'),
]

# ── 金额提取模式 ────────────────────────────────────────
# 住宿发票：不含税金额
_HOTEL_AMOUNT_PATTERNS = [
    r'(?:金额|合计)(?:[（(]不含税[）)])?\s*[:：]?\s*[¥￥]?\s*([\d.]+)',
    r'不含税[金额]?\s*[:：]?\s*[¥￥]?\s*([\d.]+)',
]
# 住宿发票：税额/税金
_HOTEL_TAX_PATTERNS = [
    r'(?:税额|税金|税\s*额|税\s*金)\s*[:：]?\s*[¥￥]?\s*([\d.]+)',
    r'(?:税额|税金)[（(]小写[）)]\s*[:：]?\s*[¥￥]?\s*([\d.]+)',
]


def _determine_category(text: str) -> str:
    """根据文本内容判断费用类别
    
    返回: 'transport' | 'hotel' | 'unknown'

    注意：交通关键字优先于住宿关键字，因为滴滴行程单的目的地可能包含"酒店"，
    但行程单本身是交通费用，不应被误判为住宿。
    """
    # 优先检查交通关键字（滴滴行程单可能含"酒店"目的地，但本质是交通费）
    for kw in _TRANSPORT_KEYWORDS:
        if kw in text:
            return 'transport'
    # 再检查住宿关键字
    for kw in _HOTEL_KEYWORDS:
        if kw in text:
            return 'hotel'
    return 'unknown'


def _determine_category_from_filename(filename: str) -> str | None:
    """根据文件名判断费用类别（fallback）

    输出文件名后缀携带类型信息：
        *高铁票.pdf → transport
    """
    if '高铁票' in filename:
        return 'transport'
    if '行程单' in filename:
        return 'transport'
    return None


def _extract_date(text: str, category: str = 'unknown') -> str | None:
    """从发票文本中提取日期

    根据类别按优先级匹配日期字段：
    - transport: 乘车日期 > 行程日期 > 发车时间日期 > 开票日期 > 通用日期
    - hotel:     入住日期 > 开票日期 > 通用日期
    - unknown:   首个特定前缀日期 > 通用日期
    返回 'YYYY-MM-DD' 格式字符串，失败返回 None。
    """

    def _fmt(m: re.Match) -> str:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

    if category == 'transport':
        # ① 优先匹配行程日期（乘车、行程），不匹配开票/发票日期
        for prefix in ('乘车', '行程'):
            m = re.search(rf'{prefix}日期\s*[:：]?\s*(\d{{4}})[年/-](\d{{1,2}})[月/-](\d{{1,2}})[日]?', text)
            if m:
                return _fmt(m)
        # ② 高铁票发车时间格式：2026年07月22日 19:00开
        m = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日\s*\d{2}:\d{2}开', text)
        if m:
            return _fmt(m)
        # ③ fallback 到开票日期
        m = re.search(r'(?:开票|开具)日期\s*[:：]?\s*(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})[日]?', text)
        if m:
            return _fmt(m)

    elif category == 'hotel':
        # ① 优先匹配入住日期
        m = re.search(r'入住日期\s*[:：]?\s*(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})[日]?', text)
        if m:
            return _fmt(m)
        # ② fallback 到开票日期
        m = re.search(r'(?:开票|开具)日期\s*[:：]?\s*(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})[日]?', text)
        if m:
            return _fmt(m)

    # unknown 或以上 fallback 均未命中 → 所有特定前缀日期，取首个
    matches = list(re.finditer(_DATE_PATTERNS[0], text))
    if matches:
        return _fmt(matches[0])

    # 最终 fallback：通用日期，取最后一个
    matches = list(re.finditer(_DATE_PATTERNS[1], text))
    if matches:
        return _fmt(matches[-1])

    return None


def _extract_amount_from_filename(filename: str) -> float | None:
    """从文件名中提取金额（价税合计）

    文件名格式: {invoice_no}-{amount}.pdf 或 {invoice_no}-{amount}行程单.pdf
    """
    m = re.match(r'^[^-]+-(\d+\.\d{2})', filename)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _extract_hotel_amounts(text: str) -> tuple[float | None, float | None]:
    """从住宿发票文本中提取不含税金额和税金

    返回: (不含税金额, 税金)
    """
    amount = None
    tax = None

    # ① 优先匹配"合计 ¥613.21 ¥36.79"双金额格式（含税拆分）
    m = re.search(r'合计\s*[¥￥]?\s*(\d+\.\d{2})\s*[¥￥]?\s*(\d+\.\d{2})', text)
    if m:
        try:
            amount = float(m.group(1))
            tax = float(m.group(2))
            # 如果金额和税金都提取成功，直接返回
            if amount is not None and tax is not None:
                return amount, tax
        except ValueError:
            pass

    # ② 单独提取不含税金额
    for p in _HOTEL_AMOUNT_PATTERNS:
        m = re.search(p, text)
        if m:
            try:
                amount = float(m.group(1))
            except ValueError:
                pass
            break

    # ③ 单独提取税额
    for p in _HOTEL_TAX_PATTERNS:
        m = re.search(p, text)
        if m:
            try:
                tax = float(m.group(1))
            except ValueError:
                pass
            break

    return amount, tax


def _parse_didi_trip_details(raw_text: str) -> list[dict] | None:
    """解析滴滴行程单中的逐笔行程明细

    从原始文本（保留空白）中提取每笔行程的日期、时间、城市、路线与金额。
    返回: [{'date', 'time', 'city', 'route', 'mileage', 'amount'}, ...] 或 None
    """
    # 提取年份（从行程起止日期）
    ym = re.search(r'行程起止日期.*?(\d{4})-\d{2}-\d{2}', raw_text)
    if not ym:
        return None
    year = ym.group(1)

    trips = []
    for line in raw_text.split('\n'):
        line = line.strip()
        if not line:
            continue
        # 行程行: 序号 车型 MM-DD HH:MM 星期 城市 起点|终点 ... 里程 金额
        m = re.match(
            r'^\d+\s+\S+\s+(\d{2})-(\d{2})\s+(\d{2}:\d{2})\s+\S+\s+(.+?)\s+(\d+\.\d+)\s+(\d+\.\d+)\s*$',
            line,
        )
        if not m:
            continue
        mm, dd, hhmm, route, mileage, amount = m.groups()
        city = route.split()[0] if route.split() else ''
        try:
            amount = float(amount)
            mileage = float(mileage)
        except ValueError:
            continue
        if amount > 0:
            trips.append({
                'date': f'{year}-{mm}-{dd}',
                'time': hhmm,
                'city': city,
                'route': route,
                'mileage': mileage,
                'amount': amount,
            })

    return trips if trips else None


def _parse_invoice(output_dir: str, filename: str, processor) -> list[dict] | None:
    """解析单个发票 PDF，返回结构化数据列表

    对于行程单（含逐笔明细），返回多行数据，每行一个日期一条金额；
    对于普通发票，返回单行数据。

    返回: [{
        'date': 'YYYY-MM-DD',
        'transport_amount': float | None,
        'hotel_amount': float | None,
        'hotel_tax': float | None,
        'category': 'transport' | 'hotel' | 'unknown',
    }, ...] 或 None
    """
    file_path = os.path.join(output_dir, filename)
    if not os.path.isfile(file_path):
        return None

    text, error_type = processor.extract_pdf_text_with_error(file_path)
    if not text:
        return None

    # 判断类别
    category = _determine_category(text)
    if category == 'unknown':
        category = _determine_category_from_filename(filename) or 'unknown'

    # ── 行程单：按天拆分逐笔行程 ──
    if '行程单' in filename and category == 'transport':
        # 用原始文本（保留空白）解析行程明细
        raw_text, _ = processor._extract_raw_text(file_path)
        if raw_text:
            trips = _parse_didi_trip_details(raw_text)
            if trips:
                return [
                    {
                        'date': t['date'],
                        'transport_amount': t['amount'],
                        'hotel_amount': None,
                        'hotel_tax': None,
                        'category': 'transport',
                        'time': t.get('time'),
                        'city': t.get('city'),
                        'route': t.get('route'),
                    }
                    for t in trips
                ]

    # ── 普通发票（含高铁票、住宿发票等） ──
    date = _extract_date(text, category)

    transport_amount = None
    hotel_amount = None
    hotel_tax = None
    extra: dict = {}

    if category == 'hotel':
        hotel_amount, hotel_tax = _extract_hotel_amounts(text)
    elif category == 'transport':
        transport_amount = _extract_amount_from_filename(filename)
        # 高铁票发车时间：2026年07月22日 19:00开
        m = re.search(r'\d{4}年\d{1,2}月\d{1,2}日\s*(\d{1,2}):(\d{2})\s*开', text)
        if m:
            extra['time'] = f'{int(m.group(1)):02d}:{m.group(2)}'

    if category == 'unknown':
        transport_amount = _extract_amount_from_filename(filename)

    if date is None:
        return None

    return [{
        'date': date,
        'transport_amount': transport_amount,
        'hotel_amount': hotel_amount,
        'hotel_tax': hotel_tax,
        'category': category,
        **extra,
    }]


# ── Excel 样式常量 ──────────────────────────────────────
_HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
_HEADER_FONT = Font(name='Microsoft YaHei UI', size=11, bold=True, color='FFFFFF')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

_DATA_FONT = Font(name='Microsoft YaHei UI', size=10)
_DATA_ALIGN = Alignment(horizontal='center', vertical='center')
_NUM_ALIGN = Alignment(horizontal='right', vertical='center')

_TOTAL_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
_TOTAL_FONT = Font(name='Microsoft YaHei UI', size=10, bold=True)

_THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

_NOTE_FONT = Font(name='Microsoft YaHei UI', size=9, color='666666')


def _build_workbook(rows: list[dict]) -> openpyxl.Workbook:
    """构建 Excel 工作簿

    rows: 按日期排序的字典列表，每项含:
        date, transport_amount, hotel_amount, hotel_tax
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '费用汇总'

    # 列宽
    ws.column_dimensions['A'].width = 14   # 日期
    ws.column_dimensions['B'].width = 16   # 交通费
    ws.column_dimensions['C'].width = 16   # 住宿费
    ws.column_dimensions['D'].width = 16   # 住宿费税金

    # ── 表头 ──
    headers = ['日期', '交通费', '住宿费', '住宿费税金']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # ── 数据行 ──
    total_transport = 0.0
    total_hotel = 0.0
    total_hotel_tax = 0.0

    for i, row_data in enumerate(rows, 2):
        date_str = row_data['date']
        trans_amt = row_data['transport_amount'] or 0.0
        hotel_amt = row_data['hotel_amount'] or 0.0
        hotel_tax = row_data['hotel_tax'] or 0.0

        total_transport += trans_amt
        total_hotel += hotel_amt
        total_hotel_tax += hotel_tax

        # 日期
        c = ws.cell(row=i, column=1, value=date_str)
        c.font = _DATA_FONT
        c.alignment = _DATA_ALIGN
        c.border = _THIN_BORDER

        # 交通费
        c = ws.cell(row=i, column=2, value=trans_amt if trans_amt > 0 else None)
        c.font = _DATA_FONT
        c.alignment = _NUM_ALIGN
        c.border = _THIN_BORDER
        c.number_format = '#,##0.00'

        # 住宿费
        c = ws.cell(row=i, column=3, value=hotel_amt if hotel_amt > 0 else None)
        c.font = _DATA_FONT
        c.alignment = _NUM_ALIGN
        c.border = _THIN_BORDER
        c.number_format = '#,##0.00'

        # 住宿费税金
        c = ws.cell(row=i, column=4, value=hotel_tax if hotel_tax > 0 else None)
        c.font = _DATA_FONT
        c.alignment = _NUM_ALIGN
        c.border = _THIN_BORDER
        c.number_format = '#,##0.00'

    # ── 合计行 ──
    total_row = len(rows) + 2
    c = ws.cell(row=total_row, column=1, value='合计')
    c.fill = _TOTAL_FILL
    c.font = _TOTAL_FONT
    c.alignment = _DATA_ALIGN
    c.border = _THIN_BORDER

    for col, val in [(2, total_transport), (3, total_hotel), (4, total_hotel_tax)]:
        c = ws.cell(row=total_row, column=col, value=val)
        c.fill = _TOTAL_FILL
        c.font = _TOTAL_FONT
        c.alignment = _NUM_ALIGN
        c.border = _THIN_BORDER
        c.number_format = '#,##0.00'

    # ── 备注 ──
    note_row = total_row + 2
    notes = [
        '备注：',
        '1. 交通费 = 价税合计（高铁按乘车日期、打车按行程日期、通行费按开票日期归集）',
        '2. 住宿费 = 不含税金额，住宿费税金 = 税额（按入住日期归集，跨天归到入住日）',
        f'3. 生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
    ]
    for j, note in enumerate(notes):
        c = ws.cell(row=note_row + j, column=1, value=note)
        c.font = _NOTE_FONT

    return wb


def generate_expense_summary(output_dir: str, processor) -> str | None:
    """生成费用汇总 Excel 文件

    遍历输出目录中的 PDF，提取日期、类别、金额，按日期归集后生成 Excel。

    Args:
        output_dir: 处理后的输出目录（包含 PDF 文件）
        processor: InvoiceProcessor 实例（用于文本提取缓存）

    Returns:
        Excel 文件路径，失败返回 None
    """
    try:
        pdf_files = []
        for f in os.listdir(output_dir):
            if not f.lower().endswith('.pdf'):
                continue
            # 跳过子目录中的文件
            if not os.path.isfile(os.path.join(output_dir, f)):
                continue
            # 跳过合并结果（合并后的 PDF 包含所有发票，单独处理会导致重复）
            if f == '合并结果.pdf':
                continue
            pdf_files.append(f)

        # 去重：同号发票同时存在 *.pdf 和 *行程单.pdf 时，只保留行程单
        # 行程单包含详细行程日期，比正式发票更准确
        inv_set: dict[str, list[str]] = {}
        for f in pdf_files:
            inv_no = f.split('-')[0]  # 发票号在第一个 - 之前
            inv_set.setdefault(inv_no, []).append(f)

        skip_set: set[str] = set()
        for inv_no, names in inv_set.items():
            has_trip = any('行程单' in n for n in names)
            if has_trip and len(names) > 1:
                # 有行程单版本，跳过非行程单版本
                for n in names:
                    if '行程单' not in n:
                        skip_set.add(n)

        if not pdf_files:
            logger.warning('输出目录中无 PDF 文件，无法生成费用汇总')
            return None

        # 逐文件解析
        data_by_date: dict[str, dict] = defaultdict(
            lambda: {'transport_amount': 0.0, 'hotel_amount': 0.0, 'hotel_tax': 0.0}
        )
        parsed_count = 0
        skipped_count = 0

        for filename in pdf_files:
            if filename in skip_set:
                skipped_count += 1
                continue

            results = _parse_invoice(output_dir, filename, processor)
            if results is None:
                continue

            parsed_count += 1

            for result in results:
                date = result['date']
                if result['transport_amount'] is not None:
                    data_by_date[date]['transport_amount'] += result['transport_amount']
                if result['hotel_amount'] is not None:
                    data_by_date[date]['hotel_amount'] += result['hotel_amount']
                if result['hotel_tax'] is not None:
                    data_by_date[date]['hotel_tax'] += result['hotel_tax']

        if not data_by_date:
            logger.warning('未能从任何 PDF 中提取到有效数据')
            return None

        # 按日期排序
        sorted_dates = sorted(data_by_date.keys())
        rows = [
            {
                'date': d,
                'transport_amount': data_by_date[d]['transport_amount'],
                'hotel_amount': data_by_date[d]['hotel_amount'],
                'hotel_tax': data_by_date[d]['hotel_tax'],
            }
            for d in sorted_dates
        ]

        # 生成 Excel
        wb = _build_workbook(rows)
        excel_path = os.path.join(output_dir, '费用汇总.xlsx')
        wb.save(excel_path)
        logger.info(
            f'费用汇总已生成: {excel_path}'
            f'（{parsed_count} 个文件，{len(rows)} 个日期'
            f'{f"，跳过 {skipped_count} 个重复发票" if skipped_count else ""}）'
        )
        return excel_path

    except Exception as e:
        logger.error(f'费用汇总生成失败: {e}', exc_info=True)
        return None