"""AI 审核模块（DeepSeek API，OpenAI 兼容接口）

对处理后的发票/行程数据进行 AI 审核，**重点是发现金额错误**（发票号/价税合计/税额/
税率/行程合计一致性等），行程只做简易核对（避免填错），不做复杂的时间/城市推理。
- 仅使用标准库（urllib），无新增依赖
- 审核结果只作提示，不阻断处理流程
"""
import json
import logging
import os
import urllib.error
import urllib.request

logger = logging.getLogger(__name__)

DEFAULT_API_BASE = 'https://api.deepseek.com'
DEFAULT_MODEL = 'DeepSeek-V4-Flash'

_SYSTEM_PROMPT = (
    '你是一名严谨的财务发票审核员。下面是某次出差产生的发票与行程数据（JSON），'
    '每项含文件名与解析出的行数据（日期/类别/金额/税额等）。请审核，**重点是金额错误**：\n'
    '1. 金额/票据异常：发票号码、价税合计、金额、税额明显不合理；同号发票金额不一致；'
    '行程合计与对应发票价税合计不一致；税率异常（如住宿不是 3%/6%/9%）；金额异常高；\n'
    '2. 行程简易核对（只做简单检查，避免填错）：行程日期与开票日期明显矛盾、'
    '同一天同一路线重复、行程单与发票不配套；不需要做复杂的时间/城市推理；\n'
    '3. 重复或可疑费用：疑似重复报销、同一天多笔同类费用异常。\n'
    '只列出确实有问题的项；没有问题就输出空数组。\n'
    '严格只输出 JSON 数组，不要输出任何其他文字。'
    '每项格式：{"file": "文件名", "type": "extract|conflict|duplicate|other", '
    '"issue": "问题描述", "suggestion": "建议"}'
)


def build_prompt(records: list[dict]) -> str:
    """构造审核提示词（把解析数据序列化为 JSON 注入）"""
    data = json.dumps(records, ensure_ascii=False, indent=1)
    return f"{_SYSTEM_PROMPT}\n\n数据：\n{data}"


def parse_findings(content: str) -> list[dict]:
    """从模型输出中宽容解析 JSON 数组（容忍 markdown 围栏、前后缀噪声）"""
    if not content:
        return []
    text = content.strip()
    # 去掉可能的 ```json ... ``` 围栏
    if text.startswith('```'):
        text = text.strip('`').strip()
        if text.startswith('json'):
            text = text[4:].strip()
    # 截取第一个 [ 到最后一个 ]
    start, end = text.find('['), text.rfind(']')
    if start == -1 or end == -1 or end <= start:
        return []
    try:
        data = json.loads(text[start:end + 1])
    except (ValueError, TypeError):
        return []
    return [d for d in data if isinstance(d, dict)] if isinstance(data, list) else []


def audit_records(records: list[dict], api_key: str,
                  api_base: str = DEFAULT_API_BASE, model: str = DEFAULT_MODEL,
                  timeout: int = 60) -> list[dict]:
    """调用 DeepSeek 审核，返回问题列表

    Raises:
        RuntimeError: API 调用失败或响应格式异常（由调用方记录日志，不中断主流程）
    """
    if not records:
        return []
    if not api_key:
        raise ValueError('未配置 DeepSeek API Key（config.ini [ai] 段）')

    payload = {
        'model': model,
        'messages': [
            {'role': 'user', 'content': build_prompt(records)},
        ],
        'temperature': 0,
        'stream': False,
    }
    req = urllib.request.Request(
        f"{api_base.rstrip('/')}/chat/completions",
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {api_key}',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        detail = e.read().decode('utf-8', errors='replace')[:200]
        raise RuntimeError(f'DeepSeek HTTP {e.code}: {detail}') from e
    except urllib.error.URLError as e:
        raise RuntimeError(f'DeepSeek 网络错误: {e.reason}') from e

    try:
        content = body['choices'][0]['message']['content']
    except (KeyError, IndexError, TypeError):
        raise RuntimeError(f'DeepSeek 响应格式异常: {str(body)[:200]}') from None
    return parse_findings(content)


def write_audit_report(output_dir: str, findings: list[dict]) -> str | None:
    """把审核结果回填到 费用汇总.xlsx 的「审核报告」工作表

    findings 每项可含 source（如「本地规则」/「AI 审核」）。
    返回 Excel 路径；费用汇总不存在或写入失败时返回 None。
    """
    xlsx_path = os.path.join(output_dir, '费用汇总.xlsx')
    if not os.path.isfile(xlsx_path):
        return None
    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        if '审核报告' in wb.sheetnames:
            del wb['审核报告']
        ws = wb.create_sheet('审核报告')
        ws.append(['来源', '文件', '类型', '问题', '建议'])
        for item in findings:
            ws.append([
                item.get('source', ''),
                item.get('file', ''),
                item.get('type', ''),
                item.get('issue', ''),
                item.get('suggestion', ''),
            ])
        if not findings:
            ws.append(['', '', '', '无异常', ''])
        for col, width in zip('ABCDE', (10, 30, 10, 60, 40)):
            ws.column_dimensions[col].width = width
        wb.save(xlsx_path)
        logger.info('审核报告已写入: %s', xlsx_path)
        return xlsx_path
    except Exception as e:
        logger.warning('审核报告写入 Excel 失败: %s', e)
        return None
